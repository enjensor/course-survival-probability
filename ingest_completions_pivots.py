#!/usr/bin/env python3
"""
Ingest historical Award Course Completions Pivot Tables (2021-2023) and
Enrolment Pivot Tables (2021-2023) into he_stats.db.

These files have a Pivot_BFOE sheet with field-level data per institution.
The 2024 "Perturbed" pivot tables were already ingested; this script fills
in the gap so that completions_trend and enrolment_trend have 4 years of data.

Usage:
    python3 ingest_completions_pivots.py
"""
from __future__ import annotations

import os
import re
import sqlite3
from typing import Dict, Optional

import openpyxl

DB_PATH = "he_stats.db"
DATA_DIR = "_downloads/files"

# Field name normalisation — map various header spellings to canonical names
FIELD_CANONICAL = {
    "natural and physical sciences": "Natural and Physical Sciences",
    "information technology": "Information Technology",
    "engineering and related technologies": "Engineering and Related Technologies",
    "architecture and building": "Architecture and Building",
    "agriculture environmental and related studies": "Agriculture, Environmental and Related Studies",
    "agriculture, environmental and related studies": "Agriculture, Environmental and Related Studies",
    "health": "Health",
    "education": "Education",
    "management and commerce": "Management and Commerce",
    "society and culture": "Society and Culture",
    "creative arts": "Creative Arts",
    "food hospitality and personal services": "Food, Hospitality and Personal Services",
    "food, hospitality and personal services": "Food, Hospitality and Personal Services",
    "mixed field programmes": "Mixed Field Programmes",
    "non-award courses": "Non-Award Courses",
}

# Files to ingest
COMPLETIONS_PIVOTS = [
    ("2021 Pivot Table Award Course Completions-ACC.xlsx", 2021),
    ("2022 Award Course Completions Pivot Table.xlsx", 2022),
    ("2023 Award Course Completions Pivot Table.xlsx", 2023),
]

ENROLMENT_PIVOTS = [
    ("2021 Pivot Table Student Enrolment.xlsx", 2021),
    ("2022 Student Enrolment Pivot Table.xlsx", 2022),
    ("2023 Student Enrolment Pivot Table_updated02.xlsx", 2023),
]

# Skip patterns for institution names
SKIP_PATTERNS = [
    "national total", "table a provider", "table b provider",
    "table c provider", "non-university higher education",
    "total:", "total ", "grand total",
]


def normalise_inst_name(raw: str) -> str:
    """Normalise institution name for matching."""
    s = raw.strip()
    s = re.sub(r"\(\d+\.\d+\)", "", s)  # footnote refs
    s = re.sub(r"\(\d{4}\)", "", s)  # inst codes
    s = re.sub(r"\s+", " ", s).strip()
    s = s.rstrip(",.")
    return s


def resolve_institution(conn: sqlite3.Connection, raw_name: str, state: str = "") -> Optional[int]:
    """Find institution ID by name, checking aliases too."""
    norm = normalise_inst_name(raw_name)
    if not norm:
        return None

    for pat in SKIP_PATTERNS:
        if pat in norm.lower():
            return None

    # Direct match
    row = conn.execute(
        "SELECT id FROM institutions WHERE name = ?", (norm,)
    ).fetchone()
    if row:
        return row[0]

    # Alias match
    row = conn.execute(
        "SELECT institution_id FROM institution_aliases WHERE alias = ?", (norm,)
    ).fetchone()
    if row:
        return row[0]

    # Fuzzy: try "The University of X" -> "University of X" and vice versa
    if norm.startswith("The "):
        alt = norm[4:]
    else:
        alt = "The " + norm

    row = conn.execute(
        "SELECT id FROM institutions WHERE name = ?", (alt,)
    ).fetchone()
    if row:
        return row[0]

    row = conn.execute(
        "SELECT institution_id FROM institution_aliases WHERE alias = ?", (alt,)
    ).fetchone()
    if row:
        return row[0]

    print(f"  [WARN] Could not resolve institution: {norm!r} (state={state!r})")
    return None


def resolve_field(conn: sqlite3.Connection, raw_name: str) -> Optional[int]:
    """Map a raw field column header to a field_id."""
    key = raw_name.strip().lower()
    # Remove leading/trailing spaces that appear in some pivot headers
    key = re.sub(r"\s+", " ", key).strip()

    # Try canonical mapping
    if key in FIELD_CANONICAL:
        canonical = FIELD_CANONICAL[key]
        row = conn.execute(
            "SELECT id FROM fields_of_education WHERE broad_field = ?", (canonical,)
        ).fetchone()
        if row:
            return row[0]

    # Try direct DB match
    row = conn.execute(
        "SELECT id FROM fields_of_education WHERE LOWER(broad_field) = ?", (key,)
    ).fetchone()
    if row:
        return row[0]

    # Partial match
    row = conn.execute(
        "SELECT id FROM fields_of_education WHERE LOWER(broad_field) LIKE ?",
        (key[:20] + "%",)
    ).fetchone()
    if row:
        return row[0]

    return None


def parse_bfoe_sheet(conn: sqlite3.Connection, filepath: str, data_year: int,
                     is_completions: bool) -> int:
    """Parse the Pivot_BFOE sheet from a pivot table workbook."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find BFOE sheet
    target_sheet = None
    for sn in wb.sheetnames:
        if "bfoe" in sn.lower():
            target_sheet = sn
            break

    if target_sheet is None:
        print(f"  [WARN] No BFOE sheet found in {os.path.basename(filepath)}")
        wb.close()
        return 0

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row with 'State', 'Institution', and field columns
    header_row_idx = None
    state_col = None
    inst_col = None
    field_cols: Dict[int, str] = {}

    for i, row in enumerate(rows):
        if not row:
            continue
        row_strs = [str(v).strip() if v is not None else "" for v in row]

        has_state = any(s.lower() == "state" for s in row_strs)
        has_inst = any("institution" in s.lower() for s in row_strs)

        if has_state and has_inst:
            header_row_idx = i
            for j, s in enumerate(row_strs):
                sl = s.lower().strip()
                if sl == "state":
                    state_col = j
                elif "institution" in sl:
                    inst_col = j
                else:
                    # Try to match as field name
                    field_id = resolve_field(conn, s)
                    if field_id is not None:
                        field_cols[j] = field_id
            break

    if header_row_idx is None or not field_cols:
        print(f"  [WARN] Could not find header row in {os.path.basename(filepath)}")
        wb.close()
        return 0

    if state_col is None:
        state_col = 0
    if inst_col is None:
        inst_col = 1

    # Also check if the Year is specified in the sheet (as a filter row)
    detected_year = None
    for i in range(header_row_idx):
        row = rows[i]
        if row and row[0] is not None and str(row[0]).strip().lower() == "year":
            val = row[1]
            if isinstance(val, (int, float)) and 2000 < val < 2030:
                detected_year = int(val)
                break
    if detected_year and detected_year != data_year:
        print(f"  [INFO] Detected year {detected_year} in sheet (expected {data_year})")
        data_year = detected_year

    fname = os.path.basename(filepath)
    total_rows = 0
    current_state = ""

    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if not row:
            continue

        # Update state
        sv = row[state_col] if state_col < len(row) else None
        if sv is not None and str(sv).strip():
            current_state = str(sv).strip()

        # Get institution
        inst_raw = row[inst_col] if inst_col < len(row) else None
        if inst_raw is None or not str(inst_raw).strip():
            continue
        inst_name = str(inst_raw).strip()

        # Skip footnotes
        if inst_name.startswith("(") or inst_name.startswith("Grand Total"):
            continue

        inst_id = resolve_institution(conn, inst_name, state=current_state)
        if inst_id is None:
            continue

        # Extract field values
        for col_idx, field_id in field_cols.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                headcount = int(round(float(val)))
            except (ValueError, TypeError):
                continue
            if headcount == 0:
                continue

            if is_completions:
                try:
                    conn.execute(
                        "INSERT INTO completions "
                        "(institution_id, year, field_id, course_level, headcount, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (inst_id, data_year, field_id, "All", headcount, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass
            else:
                try:
                    conn.execute(
                        "INSERT INTO enrolments "
                        "(institution_id, year, field_id, course_level, student_type, commencing, headcount, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                        (inst_id, data_year, field_id, "All", "all", 0, headcount, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass

    wb.close()
    return total_rows


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row

    total = 0

    # Ingest completions pivots
    print("=== Ingesting Award Course Completions Pivot Tables ===")
    for fname, year in COMPLETIONS_PIVOTS:
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  [SKIP] File not found: {fname}")
            continue

        # Check if already ingested
        row = conn.execute(
            "SELECT 1 FROM ingested_files WHERE filename = ?", (fname,)
        ).fetchone()
        if row:
            print(f"  [SKIP] Already ingested: {fname}")
            continue

        print(f"  Parsing: {fname}")
        rows = parse_bfoe_sheet(conn, fpath, year, is_completions=True)
        conn.execute(
            "INSERT OR IGNORE INTO ingested_files (filename, file_path, row_count, section, data_year) "
            "VALUES (?, ?, ?, ?, ?)",
            (fname, fpath, rows, "pivot-completions", str(year)),
        )
        conn.commit()
        total += rows
        print(f"    -> {rows} rows inserted")

    # Ingest enrolment pivots
    print("\n=== Ingesting Student Enrolment Pivot Tables ===")
    for fname, year in ENROLMENT_PIVOTS:
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  [SKIP] File not found: {fname}")
            continue

        row = conn.execute(
            "SELECT 1 FROM ingested_files WHERE filename = ?", (fname,)
        ).fetchone()
        if row:
            print(f"  [SKIP] Already ingested: {fname}")
            continue

        print(f"  Parsing: {fname}")
        rows = parse_bfoe_sheet(conn, fpath, year, is_completions=False)
        conn.execute(
            "INSERT OR IGNORE INTO ingested_files (filename, file_path, row_count, section, data_year) "
            "VALUES (?, ?, ?, ?, ?)",
            (fname, fpath, rows, "pivot-enrolments", str(year)),
        )
        conn.commit()
        total += rows
        print(f"    -> {rows} inserted")

    # Summary
    print(f"\n{'='*60}")
    print(f"Total rows inserted: {total}")

    # Check coverage
    for table, label in [("completions", "Completions"), ("enrolments", "Enrolments")]:
        print(f"\n{label} coverage (field-level):")
        cur = conn.execute(f"""
            SELECT year, COUNT(DISTINCT institution_id) as inst_count, COUNT(*) as rows
            FROM {table}
            WHERE field_id IS NOT NULL
            GROUP BY year
            ORDER BY year
        """)
        for r in cur:
            print(f"  {r['year']}: {r['inst_count']} institutions, {r['rows']} rows")

    conn.close()


if __name__ == "__main__":
    main()
