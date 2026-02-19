#!/usr/bin/env python3
"""
Ingest downloaded Higher Education Statistics files into SQLite.

Reads Excel/XLS/ODS files from the download directory, parses the structured
tables within them, normalises institution names, and loads the data into
a SQLite database designed for the Course Survival Probability engine.

Usage:
    python ingest.py --db he_stats.db --data-dir _downloads/files/
"""
from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

# ---------------------------------------------------------------------------
# Schema DDL (mirrors schema.sql)
# ---------------------------------------------------------------------------

SCHEMA_DDL = """
CREATE TABLE IF NOT EXISTS institutions (
    id          INTEGER PRIMARY KEY,
    code        TEXT UNIQUE,
    name        TEXT NOT NULL,
    state       TEXT,
    provider_type TEXT
);
CREATE TABLE IF NOT EXISTS institution_aliases (
    alias       TEXT PRIMARY KEY,
    institution_id INTEGER NOT NULL REFERENCES institutions(id)
);
CREATE TABLE IF NOT EXISTS fields_of_education (
    id          INTEGER PRIMARY KEY,
    broad_field TEXT NOT NULL UNIQUE
);
CREATE TABLE IF NOT EXISTS attrition_retention (
    id              INTEGER PRIMARY KEY,
    institution_id  INTEGER NOT NULL REFERENCES institutions(id),
    year            INTEGER NOT NULL,
    student_type    TEXT NOT NULL,
    measure         TEXT NOT NULL,
    rate            REAL,
    source_file     TEXT,
    UNIQUE(institution_id, year, student_type, measure)
);
CREATE TABLE IF NOT EXISTS completion_rates (
    id              INTEGER PRIMARY KEY,
    institution_id  INTEGER NOT NULL REFERENCES institutions(id),
    cohort_start    INTEGER NOT NULL,
    cohort_end      INTEGER NOT NULL,
    duration_years  INTEGER NOT NULL,
    completed_pct       REAL,
    still_enrolled_pct  REAL,
    dropped_out_pct     REAL,
    never_returned_pct  REAL,
    source_file     TEXT,
    UNIQUE(institution_id, cohort_start, duration_years)
);
CREATE TABLE IF NOT EXISTS enrolments (
    id              INTEGER PRIMARY KEY,
    institution_id  INTEGER NOT NULL REFERENCES institutions(id),
    year            INTEGER NOT NULL,
    field_id        INTEGER REFERENCES fields_of_education(id),
    course_level    TEXT,
    student_type    TEXT,
    commencing      INTEGER,
    headcount       INTEGER,
    eftsl           REAL,
    source_file     TEXT
);
CREATE TABLE IF NOT EXISTS completions (
    id              INTEGER PRIMARY KEY,
    institution_id  INTEGER NOT NULL REFERENCES institutions(id),
    year            INTEGER NOT NULL,
    field_id        INTEGER REFERENCES fields_of_education(id),
    course_level    TEXT,
    headcount       INTEGER,
    source_file     TEXT
);
CREATE TABLE IF NOT EXISTS ingested_files (
    id          INTEGER PRIMARY KEY,
    filename    TEXT UNIQUE NOT NULL,
    file_path   TEXT,
    ingested_at TEXT DEFAULT (datetime('now')),
    row_count   INTEGER,
    section     TEXT,
    data_year   TEXT
);
CREATE INDEX IF NOT EXISTS idx_ar_inst_year
    ON attrition_retention(institution_id, year);
CREATE INDEX IF NOT EXISTS idx_cr_inst_start
    ON completion_rates(institution_id, cohort_start);
CREATE INDEX IF NOT EXISTS idx_enr_inst_year_field
    ON enrolments(institution_id, year, field_id);
CREATE INDEX IF NOT EXISTS idx_comp_inst_year_field
    ON completions(institution_id, year, field_id);
CREATE INDEX IF NOT EXISTS idx_alias_lookup
    ON institution_aliases(alias);
CREATE TABLE IF NOT EXISTS equity_performance (
    id              INTEGER PRIMARY KEY,
    institution_id  INTEGER NOT NULL REFERENCES institutions(id),
    year            INTEGER NOT NULL,
    measure         TEXT NOT NULL,
    equity_group    TEXT NOT NULL,
    rate            REAL,
    source_file     TEXT,
    UNIQUE(institution_id, year, measure, equity_group)
);
CREATE INDEX IF NOT EXISTS idx_eq_inst
    ON equity_performance(institution_id, year);
"""

# ---------------------------------------------------------------------------
# Broad Fields of Education (ASCED)
# ---------------------------------------------------------------------------

BROAD_FIELDS = [
    "Natural and Physical Sciences",
    "Information Technology",
    "Engineering and Related Technologies",
    "Architecture and Building",
    "Agriculture, Environmental and Related Studies",
    "Health",
    "Education",
    "Management and Commerce",
    "Society and Culture",
    "Creative Arts",
    "Food, Hospitality and Personal Services",
    "Mixed Field Programmes",
    "Non-Award Courses",
]

# ---------------------------------------------------------------------------
# Institution name normalisation
# ---------------------------------------------------------------------------

# Regex to extract institution code like (3040) from name strings
INST_CODE_RE = re.compile(r"\((\d{4})\)")

# Footnote reference patterns like (1.08), (3.03) etc embedded in institution names
FOOTNOTE_RE = re.compile(r"\(\d+\.\d+\)")


def normalise_inst_name(raw: str) -> str:
    """Produce a canonical lookup key from a raw institution name string."""
    if not isinstance(raw, str):
        return ""
    s = raw.strip()
    # Remove footnote references like (1.08)
    s = FOOTNOTE_RE.sub("", s)
    # Remove institution code like (3040)
    s = INST_CODE_RE.sub("", s)
    # Normalise whitespace
    s = re.sub(r"\s+", " ", s).strip()
    # Strip trailing commas or dots
    s = s.rstrip(",.")
    return s


def extract_inst_code(raw: str) -> Optional[str]:
    """Extract the 4-digit institution code from a raw name, if present."""
    if not isinstance(raw, str):
        return None
    m = INST_CODE_RE.search(raw)
    return m.group(1) if m else None


# Known aliases mapping variant names -> canonical name
# This handles name changes and inconsistencies across 20 years of data
KNOWN_ALIASES: Dict[str, str] = {
    "University of Technology, Sydney": "University of Technology Sydney",
    "University of Ballarat": "Federation University Australia",
    "Australian Defence Force Academy": "University of New South Wales",
    "Batchelor Institute of Indigenous Tertiary Education": "Charles Darwin University",
    "Australian Maritime College": "University of Tasmania",
    "Avondale College of Higher Education": "Avondale University",
    "Avondale College": "Avondale University",
    "MCD University of Divinity": "University of Divinity",
    "University of Sydney": "The University of Sydney",
    "University of Melbourne": "The University of Melbourne",
    "University of Queensland": "The University of Queensland",
    "University of Adelaide": "The University of Adelaide",
    "University of Western Australia": "The University of Western Australia",
    "University of Newcastle": "The University of Newcastle",
    "University of New England": "The University of New England",
}


class InstitutionRegistry:
    """Manages institution lookup, creation, and alias resolution."""

    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self._cache: Dict[str, int] = {}  # normalised name -> id
        self._load_existing()

    def _load_existing(self):
        """Load existing institutions and aliases into memory cache."""
        for row in self.conn.execute("SELECT id, name FROM institutions"):
            self._cache[normalise_inst_name(row[1])] = row[0]
        for row in self.conn.execute("SELECT alias, institution_id FROM institution_aliases"):
            self._cache[row[0]] = row[1]

    def resolve(self, raw_name: str, state: str = "",
                provider_type: str = "") -> Optional[int]:
        """Resolve a raw institution name to an ID, creating if needed."""
        if not raw_name or not isinstance(raw_name, str):
            return None

        norm = normalise_inst_name(raw_name)
        if not norm:
            return None

        # Skip aggregate rows
        skip_patterns = [
            "National Total", "Table A Provider", "Table B Provider",
            "Table C Provider", "Non-University Higher Education",
            "Total:", "Total ", "Grand Total",
        ]
        for pat in skip_patterns:
            if pat.lower() in norm.lower():
                return None

        # Check cache
        if norm in self._cache:
            return self._cache[norm]

        # Check known aliases
        canonical = KNOWN_ALIASES.get(norm, norm)
        if canonical != norm and canonical in self._cache:
            # Register this variant as an alias
            inst_id = self._cache[canonical]
            self._cache[norm] = inst_id
            self.conn.execute(
                "INSERT OR IGNORE INTO institution_aliases (alias, institution_id) VALUES (?, ?)",
                (norm, inst_id),
            )
            return inst_id

        # Also check if canonical form is already cached
        canonical_norm = normalise_inst_name(canonical)
        if canonical_norm in self._cache:
            inst_id = self._cache[canonical_norm]
            self._cache[norm] = inst_id
            self.conn.execute(
                "INSERT OR IGNORE INTO institution_aliases (alias, institution_id) VALUES (?, ?)",
                (norm, inst_id),
            )
            return inst_id

        # Create new institution
        code = extract_inst_code(raw_name)
        # If we have a code, check if it already exists
        if code:
            row = self.conn.execute(
                "SELECT id FROM institutions WHERE code = ?", (code,)
            ).fetchone()
            if row:
                inst_id = row[0]
                self._cache[norm] = inst_id
                self.conn.execute(
                    "INSERT OR IGNORE INTO institution_aliases (alias, institution_id) VALUES (?, ?)",
                    (norm, inst_id),
                )
                return inst_id

        cur = self.conn.execute(
            "INSERT INTO institutions (code, name, state, provider_type) VALUES (?, ?, ?, ?)",
            (code, canonical_norm, state or None, provider_type or None),
        )
        inst_id = cur.lastrowid
        self._cache[norm] = inst_id
        self._cache[canonical_norm] = inst_id
        # Also register original normalised form as alias
        self.conn.execute(
            "INSERT OR IGNORE INTO institution_aliases (alias, institution_id) VALUES (?, ?)",
            (norm, inst_id),
        )
        self.conn.execute(
            "INSERT OR IGNORE INTO institution_aliases (alias, institution_id) VALUES (?, ?)",
            (canonical_norm, inst_id),
        )
        return inst_id


class FieldRegistry:
    """Manages field-of-education lookup and creation."""

    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self._cache: Dict[str, int] = {}
        self._load_existing()

    def _load_existing(self):
        for row in self.conn.execute("SELECT id, broad_field FROM fields_of_education"):
            self._cache[row[1].strip().lower()] = row[0]

    def resolve(self, field_name: str) -> Optional[int]:
        if not field_name or not isinstance(field_name, str):
            return None
        key = field_name.strip().lower()
        if key in self._cache:
            return self._cache[key]
        # Try partial matching for variant names
        for known_key, fid in self._cache.items():
            if key.startswith(known_key[:20]) or known_key.startswith(key[:20]):
                self._cache[key] = fid
                return fid
        # Create new
        cur = self.conn.execute(
            "INSERT INTO fields_of_education (broad_field) VALUES (?)",
            (field_name.strip(),),
        )
        fid = cur.lastrowid
        self._cache[key] = fid
        return fid

    def seed(self):
        """Seed with known broad fields."""
        for f in BROAD_FIELDS:
            self.resolve(f)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_float(val: Any) -> Optional[float]:
    """Convert a value to float, returning None for non-numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return None
        return float(val)
    s = str(val).strip().replace(",", "")
    if s in ("", "np", "na", "n/a", "-", "...", "< 5", "<5", "n.a.", "n.p."):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def safe_int(val: Any) -> Optional[int]:
    """Convert a value to int, returning None for non-numeric."""
    f = safe_float(val)
    if f is None:
        return None
    return int(round(f))


def read_excel_safe(filepath: str, sheet_name: Any = 0, **kwargs) -> Optional[pd.DataFrame]:
    """Read an Excel sheet with error handling for format issues."""
    try:
        return pd.read_excel(filepath, sheet_name=sheet_name, header=None, **kwargs)
    except Exception as e:
        print(f"  [WARN] Cannot read sheet '{sheet_name}' in {filepath}: {e}", file=sys.stderr)
        return None


def get_sheet_names(filepath: str) -> List[str]:
    """Get sheet names from an Excel file."""
    try:
        xls = pd.ExcelFile(filepath)
        return xls.sheet_names
    except Exception as e:
        print(f"  [WARN] Cannot open {filepath}: {e}", file=sys.stderr)
        return []


def already_ingested(conn: sqlite3.Connection, filename: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM ingested_files WHERE filename = ?", (filename,)
    ).fetchone()
    return row is not None


def record_ingestion(conn: sqlite3.Connection, filename: str, filepath: str,
                     row_count: int, section: str, data_year: str):
    conn.execute(
        "INSERT OR IGNORE INTO ingested_files (filename, file_path, row_count, section, data_year) "
        "VALUES (?, ?, ?, ?, ?)",
        (filename, filepath, row_count, section, data_year),
    )


# ---------------------------------------------------------------------------
# Parser 1: Section 15 — Attrition, Retention, Success
# ---------------------------------------------------------------------------

# Maps (sheet_index_in_data_sheets, ...) to (measure, student_type)
# Section 15 has 9 data sheets in groups of 3:
#   sheets 0,1,2 = attrition (domestic, overseas, all)
#   sheets 3,4,5 = retention (domestic, overseas, all)
#   sheets 6,7,8 = success   (domestic, overseas, all)
SECTION_15_MAP = {
    0: ("attrition", "domestic"),
    1: ("attrition", "overseas"),
    2: ("attrition", "all"),
    3: ("retention", "domestic"),
    4: ("retention", "overseas"),
    5: ("retention", "all"),
    6: ("success", "domestic"),
    7: ("success", "overseas"),
    8: ("success", "all"),
}

# Older files (2018 and before) have more sheets with different numbering.
# The sheet titles contain the measure and student type, so we parse those.
S15_MEASURE_KEYWORDS = {
    "attrition": "attrition",
    "retention": "retention",
    "success": "success",
}
S15_TYPE_KEYWORDS = {
    "domestic": "domestic",
    "overseas": "overseas",
    "all ": "all",
    "all commencing": "all",
}


def classify_s15_sheet(sheet_name: str, df: pd.DataFrame) -> Optional[Tuple[str, str]]:
    """Determine (measure, student_type) from a Section 15 sheet."""
    # First try the sheet title row (usually row 1)
    title = ""
    for i in range(min(3, len(df))):
        cell = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ""
        if len(cell) > 20:
            title = cell.lower()
            break

    if not title:
        return None

    measure = None
    student_type = None

    for keyword, m in S15_MEASURE_KEYWORDS.items():
        if keyword in title:
            measure = m
            break

    for keyword, st in S15_TYPE_KEYWORDS.items():
        if keyword in title:
            student_type = st
            break

    if measure and student_type:
        return (measure, student_type)
    return None


def parse_section_15(filepath: str, conn: sqlite3.Connection,
                     registry: InstitutionRegistry) -> int:
    """Parse a Section 15 file and insert attrition/retention/success rates."""
    fname = os.path.basename(filepath)
    sheets = get_sheet_names(filepath)
    if not sheets:
        return 0

    # Filter to data sheets (skip Contents, Explanatory notes)
    data_sheets = [s for s in sheets if s.lower() not in ("contents", "explanatory notes")]

    total_rows = 0

    for sheet_name in data_sheets:
        df = read_excel_safe(filepath, sheet_name=sheet_name)
        if df is None or df.empty:
            continue

        classification = classify_s15_sheet(sheet_name, df)
        if classification is None:
            continue
        measure, student_type = classification

        # Find the header row — contains year columns
        header_row = None
        year_cols: Dict[int, int] = {}  # col_index -> year

        for i in range(min(8, len(df))):
            row_vals = df.iloc[i].tolist()
            years_found = {}
            for j, v in enumerate(row_vals):
                yr = safe_float(v)
                if yr and 1990 < yr < 2030:
                    years_found[j] = int(yr)
            if len(years_found) >= 3:
                header_row = i
                year_cols = years_found
                break

        if header_row is None:
            continue

        # Find institution name column (usually col 0 or 1)
        # And state column
        state_col = None
        inst_col = None
        for j in range(min(3, df.shape[1])):
            sample = str(df.iloc[header_row, j]).lower().strip() if pd.notna(df.iloc[header_row, j]) else ""
            if "state" in sample:
                state_col = j
            elif "institution" in sample or "higher education" in sample:
                inst_col = j

        # If no explicit columns found, assume col 0=state, col 1=institution
        if inst_col is None:
            inst_col = 1 if df.shape[1] > 1 else 0
        if state_col is None and inst_col > 0:
            state_col = 0

        # Parse data rows
        current_state = ""
        for i in range(header_row + 1, len(df)):
            # Update state from state column
            if state_col is not None:
                sv = df.iloc[i, state_col]
                if pd.notna(sv) and str(sv).strip():
                    current_state = str(sv).strip()

            # Get institution name
            inst_raw = df.iloc[i, inst_col]
            if pd.isna(inst_raw) or not str(inst_raw).strip():
                continue

            inst_name = str(inst_raw).strip()

            # Skip aggregate rows and footnotes
            if inst_name.startswith("(") or inst_name.lower().startswith("note"):
                continue

            inst_id = registry.resolve(inst_name, state=current_state)
            if inst_id is None:
                continue

            # Extract rate values for each year
            for col_idx, year in year_cols.items():
                rate = safe_float(df.iloc[i, col_idx])
                if rate is None:
                    continue
                try:
                    conn.execute(
                        "INSERT OR REPLACE INTO attrition_retention "
                        "(institution_id, year, student_type, measure, rate, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (inst_id, year, student_type, measure, rate, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass

    return total_rows


# ---------------------------------------------------------------------------
# Parser 2: Section 17 — Completion Rates
# ---------------------------------------------------------------------------

def parse_section_17(filepath: str, conn: sqlite3.Connection,
                     registry: InstitutionRegistry) -> int:
    """Parse Section 17 completion rate file."""
    fname = os.path.basename(filepath)
    sheets = get_sheet_names(filepath)
    if not sheets:
        return 0

    # Look for sheet 17.3 (per-institution) or the last data sheet
    target_sheets = [s for s in sheets if s in ("17.3", "3") or "institution" in s.lower()]
    if not target_sheets:
        # Fall back to all data sheets
        target_sheets = [s for s in sheets if s.lower() not in ("contents", "explanatory notes")]

    total_rows = 0

    for sheet_name in target_sheets:
        df = read_excel_safe(filepath, sheet_name=sheet_name)
        if df is None or df.empty:
            continue

        # Find header row with 'State', 'Institution', 'Duration', 'Timeframe'
        header_row = None
        col_map: Dict[str, int] = {}

        for i in range(min(8, len(df))):
            row_strs = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[i].tolist()]
            found = {}
            for j, s in enumerate(row_strs):
                if "state" in s or "group" in s:
                    found["state"] = j
                elif "institution" in s:
                    found["institution"] = j
                elif "duration" in s:
                    found["duration"] = j
                elif "timeframe" in s:
                    found["timeframe"] = j
                elif "completed" in s:
                    found["completed"] = j
                elif "still enrolled" in s:
                    found["still_enrolled"] = j
                elif "re-enrolled" in s or "dropped out" in s:
                    found["dropped_out"] = j
                elif "never came back" in s:
                    found["never_returned"] = j
            if len(found) >= 4:
                header_row = i
                col_map = found
                break

        if header_row is None:
            continue

        current_state = ""
        for i in range(header_row + 1, len(df)):
            # State
            if "state" in col_map:
                sv = df.iloc[i, col_map["state"]]
                if pd.notna(sv) and str(sv).strip():
                    current_state = str(sv).strip()

            # Institution
            inst_col = col_map.get("institution", col_map.get("state", 0))
            inst_raw = df.iloc[i, inst_col] if inst_col < df.shape[1] else None
            if pd.isna(inst_raw) or not str(inst_raw).strip():
                continue
            inst_name = str(inst_raw).strip()
            if inst_name.startswith("("):
                continue

            inst_id = registry.resolve(inst_name, state=current_state)
            if inst_id is None:
                continue

            # Duration
            duration_raw = df.iloc[i, col_map.get("duration", 2)] if "duration" in col_map else None
            if pd.isna(duration_raw):
                continue
            dur_str = str(duration_raw).strip().lower()
            if "four" in dur_str or "4" in dur_str:
                duration = 4
            elif "six" in dur_str or "6" in dur_str:
                duration = 6
            elif "nine" in dur_str or "9" in dur_str:
                duration = 9
            else:
                continue

            # Timeframe
            tf_raw = df.iloc[i, col_map.get("timeframe", 3)] if "timeframe" in col_map else None
            if pd.isna(tf_raw):
                continue
            tf_str = str(tf_raw).strip()
            m = re.match(r"(\d{4})-(\d{4})", tf_str)
            if not m:
                continue
            cohort_start = int(m.group(1))
            cohort_end = int(m.group(2))

            # Outcome percentages
            completed = safe_float(df.iloc[i, col_map["completed"]]) if "completed" in col_map else None
            still_enrolled = safe_float(df.iloc[i, col_map["still_enrolled"]]) if "still_enrolled" in col_map else None
            dropped_out = safe_float(df.iloc[i, col_map["dropped_out"]]) if "dropped_out" in col_map else None
            never_returned = safe_float(df.iloc[i, col_map["never_returned"]]) if "never_returned" in col_map else None

            if completed is None and still_enrolled is None:
                continue

            try:
                conn.execute(
                    "INSERT OR REPLACE INTO completion_rates "
                    "(institution_id, cohort_start, cohort_end, duration_years, "
                    "completed_pct, still_enrolled_pct, dropped_out_pct, never_returned_pct, source_file) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (inst_id, cohort_start, cohort_end, duration, completed,
                     still_enrolled, dropped_out, never_returned, fname),
                )
                total_rows += 1
            except sqlite3.IntegrityError:
                pass

    return total_rows


# ---------------------------------------------------------------------------
# Parser 3: Cohort Analysis
# ---------------------------------------------------------------------------

def parse_cohort_analysis(filepath: str, conn: sqlite3.Connection,
                          registry: InstitutionRegistry) -> int:
    """Parse standalone Cohort Analysis files (T4/T5/T6 = per-institution %)."""
    fname = os.path.basename(filepath)
    sheets = get_sheet_names(filepath)
    if not sheets:
        return 0

    # T4 = 9yr, T5 = 6yr, T6 = 4yr completion rates by institution
    duration_map = {"T4": 9, "T5": 6, "T6": 4}
    target_sheets = {s: d for s, d in duration_map.items() if s in sheets}

    if not target_sheets:
        return 0

    total_rows = 0

    for sheet_name, duration in target_sheets.items():
        df = read_excel_safe(filepath, sheet_name=sheet_name)
        if df is None or df.empty:
            continue

        # Find header row with timeframe columns (e.g. '2005-2013', '2006-2014')
        header_row = None
        timeframe_cols: Dict[int, Tuple[int, int]] = {}  # col_idx -> (start, end)

        for i in range(min(8, len(df))):
            tfs = {}
            for j in range(df.shape[1]):
                v = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ""
                m = re.match(r"(\d{4})-(\d{4})", v)
                if m:
                    tfs[j] = (int(m.group(1)), int(m.group(2)))
            if len(tfs) >= 3:
                header_row = i
                timeframe_cols = tfs
                break

        if header_row is None:
            continue

        # Parse institution rows
        current_state = ""
        for i in range(header_row + 1, len(df)):
            # Col 0 is often state, col 1 is institution
            sv = df.iloc[i, 0] if pd.notna(df.iloc[i, 0]) else ""
            if sv and isinstance(sv, str) and sv.strip() and not sv.strip().startswith("("):
                current_state = sv.strip()

            inst_raw = df.iloc[i, 1] if df.shape[1] > 1 and pd.notna(df.iloc[i, 1]) else ""
            if not inst_raw or not isinstance(inst_raw, str) or not inst_raw.strip():
                continue
            inst_name = inst_raw.strip()
            if inst_name.startswith("(") or inst_name.startswith("Note"):
                continue

            inst_id = registry.resolve(inst_name, state=current_state)
            if inst_id is None:
                continue

            for col_idx, (cs, ce) in timeframe_cols.items():
                val = safe_float(df.iloc[i, col_idx])
                if val is None:
                    continue
                try:
                    conn.execute(
                        "INSERT OR REPLACE INTO completion_rates "
                        "(institution_id, cohort_start, cohort_end, duration_years, "
                        "completed_pct, still_enrolled_pct, dropped_out_pct, never_returned_pct, source_file) "
                        "VALUES (?, ?, ?, ?, ?, NULL, NULL, NULL, ?)",
                        (inst_id, cs, ce, duration, val, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass

    return total_rows


# ---------------------------------------------------------------------------
# Parser 4: Section 2 / Section 1 — Enrolments
# ---------------------------------------------------------------------------

def parse_enrolments(filepath: str, conn: sqlite3.Connection,
                     registry: InstitutionRegistry, fields: FieldRegistry,
                     data_year: int, commencing: int) -> int:
    """
    Parse Section 2 (all students) or Section 1 (commencing students).
    Reads the field-of-education breakdown sheet (2.10/2.11 or 1.9/1.10).
    """
    fname = os.path.basename(filepath)
    sheets = get_sheet_names(filepath)
    if not sheets:
        return 0

    # Find sheets with field-of-education breakdown
    # 2024: sheets '2.10', '2.11' (all, domestic)
    # Older: may be named differently
    target_sheets = []
    for s in sheets:
        sl = s.lower().strip()
        # Sheets with field of education data are typically the later numbered ones
        # containing "field" in the title, or sheets 2.10, 2.11, 1.9, 1.10
        if sl in ("2.10", "1.9", "1.10", "2.11"):
            target_sheets.append(s)

    # If no standard sheets found, try to find by inspecting content
    if not target_sheets:
        for s in sheets:
            if s.lower() in ("contents", "explanatory notes"):
                continue
            df_probe = read_excel_safe(filepath, sheet_name=s, nrows=5)
            if df_probe is None:
                continue
            # Check if any cell contains "Field of Education"
            for i in range(min(3, len(df_probe))):
                row_text = " ".join(str(v) for v in df_probe.iloc[i].tolist() if pd.notna(v))
                if "field of education" in row_text.lower():
                    target_sheets.append(s)
                    break

    total_rows = 0

    for sheet_name in target_sheets:
        df = read_excel_safe(filepath, sheet_name=sheet_name)
        if df is None or df.empty:
            continue

        # Determine student_type from sheet title or name
        student_type = "all"
        for i in range(min(3, len(df))):
            title = str(df.iloc[i, 0]) if pd.notna(df.iloc[i, 0]) else ""
            if "domestic" in title.lower():
                student_type = "domestic"
                break

        # Find header row with field names as column headers
        header_row = None
        field_cols: Dict[int, str] = {}  # col_idx -> field_name

        for i in range(min(8, len(df))):
            fields_found = {}
            for j in range(2, df.shape[1]):
                v = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ""
                # Check if this looks like a field of education name
                for bf in BROAD_FIELDS:
                    if v and (bf.lower().startswith(v.lower()[:15]) or v.lower().startswith(bf.lower()[:15])):
                        fields_found[j] = bf
                        break
                # Also match partial names
                if v.lower().startswith("natural") or "physical science" in v.lower():
                    fields_found[j] = "Natural and Physical Sciences"
                elif v.lower().startswith("information tech"):
                    fields_found[j] = "Information Technology"
                elif "engineering" in v.lower():
                    fields_found[j] = "Engineering and Related Technologies"
                elif "architecture" in v.lower():
                    fields_found[j] = "Architecture and Building"
                elif "agriculture" in v.lower() or "environmental" in v.lower():
                    fields_found[j] = "Agriculture, Environmental and Related Studies"
                elif v.lower() == "health":
                    fields_found[j] = "Health"
                elif v.lower() == "education":
                    fields_found[j] = "Education"
                elif "management" in v.lower() or "commerce" in v.lower():
                    fields_found[j] = "Management and Commerce"
                elif "society" in v.lower() or "culture" in v.lower():
                    fields_found[j] = "Society and Culture"
                elif "creative" in v.lower():
                    fields_found[j] = "Creative Arts"
                elif "food" in v.lower() or "hospitality" in v.lower():
                    fields_found[j] = "Food, Hospitality and Personal Services"

            if len(fields_found) >= 5:
                header_row = i
                field_cols = fields_found
                break

        if header_row is None:
            continue

        # Parse data rows
        current_state = ""
        for i in range(header_row + 1, len(df)):
            sv = df.iloc[i, 0] if pd.notna(df.iloc[i, 0]) else ""
            if sv and isinstance(sv, str) and sv.strip():
                current_state = str(sv).strip()

            inst_col = 1 if df.shape[1] > 1 else 0
            inst_raw = df.iloc[i, inst_col] if pd.notna(df.iloc[i, inst_col]) else ""
            if not inst_raw or not isinstance(inst_raw, str) or not inst_raw.strip():
                continue
            inst_name = str(inst_raw).strip()
            if inst_name.startswith("(") or inst_name.startswith("Note") or inst_name.startswith("Source"):
                continue

            inst_id = registry.resolve(inst_name, state=current_state)
            if inst_id is None:
                continue

            for col_idx, field_name in field_cols.items():
                headcount = safe_int(df.iloc[i, col_idx])
                if headcount is None:
                    continue
                field_id = fields.resolve(field_name)
                try:
                    conn.execute(
                        "INSERT INTO enrolments "
                        "(institution_id, year, field_id, course_level, student_type, commencing, headcount, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                        (inst_id, data_year, field_id, "All", student_type, commencing, headcount, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass

    return total_rows


# ---------------------------------------------------------------------------
# Parser 5: Section 14 — Award Course Completions
# ---------------------------------------------------------------------------

def parse_section_14(filepath: str, conn: sqlite3.Connection,
                     registry: InstitutionRegistry, fields: FieldRegistry,
                     data_year: int) -> int:
    """Parse Section 14 award course completions. Structure similar to Section 2."""
    fname = os.path.basename(filepath)
    sheets = get_sheet_names(filepath)
    if not sheets:
        return 0

    # Find the field-of-education breakdown sheet
    target_sheets = []
    for s in sheets:
        if s.lower() in ("contents", "explanatory notes"):
            continue
        df_probe = read_excel_safe(filepath, sheet_name=s, nrows=5)
        if df_probe is None:
            continue
        for i in range(min(4, len(df_probe))):
            row_text = " ".join(str(v) for v in df_probe.iloc[i].tolist() if pd.notna(v))
            if "field of education" in row_text.lower():
                target_sheets.append(s)
                break

    total_rows = 0

    for sheet_name in target_sheets[:2]:  # Limit to first 2 matching sheets
        df = read_excel_safe(filepath, sheet_name=sheet_name)
        if df is None or df.empty:
            continue

        # Find header row with field names
        header_row = None
        field_cols: Dict[int, str] = {}

        for i in range(min(8, len(df))):
            fields_found = {}
            for j in range(2, df.shape[1]):
                v = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ""
                if v.lower().startswith("natural") or "physical science" in v.lower():
                    fields_found[j] = "Natural and Physical Sciences"
                elif v.lower().startswith("information tech"):
                    fields_found[j] = "Information Technology"
                elif "engineering" in v.lower():
                    fields_found[j] = "Engineering and Related Technologies"
                elif "architecture" in v.lower():
                    fields_found[j] = "Architecture and Building"
                elif "agriculture" in v.lower() or "environmental" in v.lower():
                    fields_found[j] = "Agriculture, Environmental and Related Studies"
                elif v.lower() == "health":
                    fields_found[j] = "Health"
                elif v.lower() == "education":
                    fields_found[j] = "Education"
                elif "management" in v.lower() or "commerce" in v.lower():
                    fields_found[j] = "Management and Commerce"
                elif "society" in v.lower() or "culture" in v.lower():
                    fields_found[j] = "Society and Culture"
                elif "creative" in v.lower():
                    fields_found[j] = "Creative Arts"
                elif "food" in v.lower() or "hospitality" in v.lower():
                    fields_found[j] = "Food, Hospitality and Personal Services"

            if len(fields_found) >= 5:
                header_row = i
                field_cols = fields_found
                break

        if header_row is None:
            continue

        current_state = ""
        for i in range(header_row + 1, len(df)):
            sv = df.iloc[i, 0] if pd.notna(df.iloc[i, 0]) else ""
            if sv and isinstance(sv, str) and sv.strip():
                current_state = str(sv).strip()

            inst_col = 1 if df.shape[1] > 1 else 0
            inst_raw = df.iloc[i, inst_col] if pd.notna(df.iloc[i, inst_col]) else ""
            if not inst_raw or not isinstance(inst_raw, str) or not inst_raw.strip():
                continue
            inst_name = str(inst_raw).strip()
            if inst_name.startswith("(") or inst_name.startswith("Note"):
                continue

            inst_id = registry.resolve(inst_name, state=current_state)
            if inst_id is None:
                continue

            for col_idx, field_name in field_cols.items():
                headcount = safe_int(df.iloc[i, col_idx])
                if headcount is None:
                    continue
                field_id = fields.resolve(field_name)
                try:
                    conn.execute(
                        "INSERT INTO completions "
                        "(institution_id, year, field_id, course_level, headcount, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (inst_id, data_year, field_id, "All", headcount, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass

    return total_rows


# ---------------------------------------------------------------------------
# Parser 6: Pivot Tables (Perturbed)
# ---------------------------------------------------------------------------

def parse_pivot_table(filepath: str, conn: sqlite3.Connection,
                      registry: InstitutionRegistry, fields: FieldRegistry,
                      is_completions: bool = False) -> int:
    """Parse perturbed pivot table files (BFOE sheet)."""
    fname = os.path.basename(filepath)
    sheets = get_sheet_names(filepath)
    if not sheets:
        return 0

    # Look for the Pivot_BFOE sheet
    target = None
    for s in sheets:
        if "bfoe" in s.lower():
            target = s
            break

    if target is None:
        return 0

    df = read_excel_safe(filepath, sheet_name=target)
    if df is None or df.empty:
        return 0

    # Find the header row with 'State', 'Institution', and field names
    header_row = None
    state_col = None
    inst_col = None
    field_cols: Dict[int, str] = {}
    year_data: Dict[int, int] = {}  # for the Pivot sheet with year columns

    for i in range(min(30, len(df))):
        row_strs = [str(v).strip() if pd.notna(v) else "" for v in df.iloc[i].tolist()]
        # Check if this looks like a header row
        if any("state" in s.lower() for s in row_strs) and any("institution" in s.lower() for s in row_strs):
            header_row = i
            for j, s in enumerate(row_strs):
                sl = s.lower()
                if "state" == sl or sl == "state":
                    state_col = j
                elif "institution" in sl:
                    inst_col = j
                else:
                    # Check field names
                    if "natural" in sl or "physical science" in sl:
                        field_cols[j] = "Natural and Physical Sciences"
                    elif "information tech" in sl:
                        field_cols[j] = "Information Technology"
                    elif "engineering" in sl:
                        field_cols[j] = "Engineering and Related Technologies"
                    elif "architecture" in sl:
                        field_cols[j] = "Architecture and Building"
                    elif "agriculture" in sl or "environmental" in sl:
                        field_cols[j] = "Agriculture, Environmental and Related Studies"
                    elif sl == "health":
                        field_cols[j] = "Health"
                    elif sl == "education":
                        field_cols[j] = "Education"
                    elif "management" in sl or "commerce" in sl:
                        field_cols[j] = "Management and Commerce"
                    elif "society" in sl or "culture" in sl:
                        field_cols[j] = "Society and Culture"
                    elif "creative" in sl:
                        field_cols[j] = "Creative Arts"
                    elif "food" in sl or "hospitality" in sl:
                        field_cols[j] = "Food, Hospitality and Personal Services"
                    elif "total" in sl and "enrolment" in sl.lower():
                        pass  # Skip total columns
            break

    if header_row is None or not field_cols:
        return 0

    if state_col is None:
        state_col = 0
    if inst_col is None:
        inst_col = 1

    # Determine the year from the filter settings above the data
    data_year = None
    for i in range(header_row):
        cell0 = str(df.iloc[i, 0]).strip().lower() if pd.notna(df.iloc[i, 0]) else ""
        cell1 = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
        if cell0 == "year":
            yr = safe_int(cell1)
            if yr and 2000 < yr < 2030:
                data_year = yr
                break

    if data_year is None:
        # Try to extract from filename
        m = re.search(r"(20\d{2})", fname)
        data_year = int(m.group(1)) if m else 2024

    total_rows = 0
    current_state = ""

    for i in range(header_row + 1, len(df)):
        sv = df.iloc[i, state_col] if pd.notna(df.iloc[i, state_col]) else ""
        if sv and isinstance(sv, str) and sv.strip():
            current_state = str(sv).strip()

        inst_raw = df.iloc[i, inst_col] if pd.notna(df.iloc[i, inst_col]) else ""
        if not inst_raw or not isinstance(inst_raw, str) or not inst_raw.strip():
            continue
        inst_name = str(inst_raw).strip()
        if inst_name.startswith("(") or inst_name.startswith("Grand Total"):
            continue

        inst_id = registry.resolve(inst_name, state=current_state)
        if inst_id is None:
            continue

        for col_idx, field_name in field_cols.items():
            headcount = safe_int(df.iloc[i, col_idx])
            if headcount is None:
                continue
            field_id = fields.resolve(field_name)

            if is_completions:
                try:
                    conn.execute(
                        "INSERT INTO completions "
                        "(institution_id, year, field_id, course_level, headcount, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (inst_id, data_year, field_id, "All", headcount, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass
            else:
                try:
                    conn.execute(
                        "INSERT INTO enrolments "
                        "(institution_id, year, field_id, course_level, student_type, commencing, headcount, source_file) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                        (inst_id, data_year, field_id, "All", "all", 0, headcount, fname),
                    )
                    total_rows += 1
                except sqlite3.IntegrityError:
                    pass

    return total_rows


# ---------------------------------------------------------------------------
# Section 16 — Equity Performance Data
# ---------------------------------------------------------------------------

# Canonical equity group keys — maps raw header substrings to our normalised names.
# We skip groups not in this map (e.g. "First Address" variants, "Undergraduate" subsets,
# combined "Regional/Remote").
_EQUITY_GROUP_MAP = {
    "All Domestic":                 "all_domestic",
    "Domestic National Total":      "all_domestic",   # 2021-2023 format
    "Non-English Speaking":         "nesb",
    "Disability":                   "disability",
    "First Nations":                "first_nations",
    "Indigenous":                   "first_nations",   # older files
    "Low SES by SA1":               "low_ses",
    "Low SES(":                     "low_ses",         # older format
    "Regional(":                    "regional",
    "Remote(":                      "remote",
}

# Groups to skip even if partially matched
_EQUITY_SKIP_PATTERNS = [
    "First Address",
    "Undergraduate",
    "Regional/Remote",
]


def _normalise_equity_group(raw: str) -> Optional[str]:
    """Map raw equity group header to canonical key, or None to skip."""
    if not raw:
        return None
    raw = raw.strip()
    # Check skip patterns first
    for skip in _EQUITY_SKIP_PATTERNS:
        if skip in raw:
            return None
    # Match against map (substring match)
    for pattern, key in _EQUITY_GROUP_MAP.items():
        if pattern in raw:
            return key
    return None


def _parse_year_value(val: Any) -> Optional[int]:
    """Parse a year value from the header row. Handles int and strings like '2016 (2011 SEIFA)'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        y = int(val)
        if 2000 <= y <= 2030:
            return y
        return None
    s = str(val).strip()
    # e.g. "2016 (2011 SEIFA)" or "2016 (2016 SEIFA)"
    m = re.match(r"^(\d{4})", s)
    if m:
        return int(m.group(1))
    return None


def _is_latest_version_column(val: Any) -> bool:
    """
    For dual-version year columns like '2016 (2011 SEIFA)' vs '2016 (2016 SEIFA)',
    return True only for the latest version. Plain integer years always return True.
    """
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True  # plain year — always keep
    s = str(val).strip()
    # Check if this is a dual-version string
    m = re.match(r"^(\d{4})\s*\((\d{4})", s)
    if m:
        data_year = int(m.group(1))
        version_year = int(m.group(2))
        # Keep only when version year matches data year (latest version)
        # or when version year is the most recent available
        # Heuristic: keep if version_year >= data_year - 1
        return version_year >= data_year - 1
    return True  # no version info — keep


def parse_section_16(
    filepath: str,
    conn: sqlite3.Connection,
    registry: "InstitutionRegistry",
) -> int:
    """
    Parse a Section 16 Equity Performance Data XLSX file.

    Reads sheets for retention (16.8 or 16.6), success (16.10 or 16.8),
    and attainment (16.13 or 16.10) rates — using the title row to
    determine the measure, not the sheet number (which varies across years).

    Returns total rows inserted.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    total_rows = 0
    fname = os.path.basename(filepath)

    for sheet_name in wb.sheetnames:
        # Accept numbered sheets: "16.8", "16.10", or just "8", "10", etc.
        stripped = sheet_name.strip()
        if not (re.match(r"^16\.\d+$", stripped) or re.match(r"^\d{1,2}$", stripped)):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Determine measure from title — scan first few rows for a title string.
        # We only want RATE sheets, not RATIO sheets.
        measure = None
        title_row_idx = None
        for i in range(min(5, len(rows))):
            cell = str(rows[i][0] or "").lower()
            # Skip ratio sheets explicitly
            if "ratio" in cell:
                break
            if "retention" in cell and "rate" in cell:
                measure = "retention"
                title_row_idx = i
                break
            elif "success" in cell and "rate" in cell:
                measure = "success"
                title_row_idx = i
                break
            elif "attainment" in cell and ("rate" in cell or "attainment" in cell):
                measure = "attainment"
                title_row_idx = i
                break
        if measure is None:
            continue

        # Find the year row — look for a row that has integer years (2009-2024) in cols 2+
        year_row_idx = None
        for i in range(title_row_idx + 1, min(title_row_idx + 5, len(rows))):
            # Check if col 2+ has year-like values
            year_count = 0
            for c in range(2, min(20, len(rows[i]))):
                y = _parse_year_value(rows[i][c])
                if y and 2000 <= y <= 2030:
                    year_count += 1
            if year_count >= 3:  # at least 3 year columns found
                year_row_idx = i
                break
        if year_row_idx is None:
            continue

        # Equity group headers are one row above the year row
        equity_header_idx = year_row_idx - 1
        data_start_idx = year_row_idx + 1

        # Build equity group -> column range map from equity header row
        equity_header = rows[equity_header_idx]
        group_start_cols: List[Tuple[int, str]] = []
        for col_idx, val in enumerate(equity_header):
            if val is not None and str(val).strip() and col_idx >= 2:
                group_start_cols.append((col_idx, str(val).strip()))

        if not group_start_cols:
            continue

        # Build col -> (equity_group_key, year) map from year row
        year_row = rows[year_row_idx]
        col_map: Dict[int, Tuple[str, int]] = {}

        for col_idx, val in enumerate(year_row):
            if col_idx < 2:
                continue
            year = _parse_year_value(val)
            if year is None:
                continue

            # Skip non-latest version columns (dual SEIFA/ASGS)
            if not _is_latest_version_column(val):
                continue

            # Find which equity group this column belongs to
            group_col = None
            group_raw = None
            for gc, gn in reversed(group_start_cols):
                if col_idx >= gc:
                    group_col = gc
                    group_raw = gn
                    break
            if group_raw is None:
                continue

            eq_key = _normalise_equity_group(group_raw)
            if eq_key is None:
                continue  # skip unwanted groups

            col_map[col_idx] = (eq_key, year)

        if not col_map:
            continue

        # Parse data rows
        current_group = None
        for row in rows[data_start_idx:]:
            col0 = row[0]
            col1 = row[1]

            # Stop at footnotes
            if col0 and str(col0).strip().startswith("("):
                break

            # Track group category (carry-forward for older files)
            if col0 and str(col0).strip():
                current_group = str(col0).strip()

            # Only process institution rows
            if current_group != "Higher Education Institution":
                continue

            if not col1 or not str(col1).strip():
                continue

            inst_name = str(col1).strip()
            inst_id = registry.resolve(inst_name)
            if inst_id is None:
                continue

            # Extract values
            for col_idx, (eq_key, year) in col_map.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None:
                    continue
                # 0 means not reported
                if isinstance(val, (int, float)) and val == 0:
                    continue
                rate = safe_float(val)
                if rate is None:
                    continue

                conn.execute(
                    """INSERT OR REPLACE INTO equity_performance
                       (institution_id, year, measure, equity_group, rate, source_file)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (inst_id, year, measure, eq_key, rate, fname),
                )
                total_rows += 1

    wb.close()
    return total_rows


# ---------------------------------------------------------------------------
# File selection & orchestration
# ---------------------------------------------------------------------------

def extract_year_from_filename(fname: str) -> Optional[int]:
    """Extract a 4-digit year from a filename."""
    m = re.search(r"(20[012]\d)", fname)
    if m:
        return int(m.group(1))
    return None


def classify_file(fname: str) -> Optional[Tuple[str, int]]:
    """
    Classify a file by its section and data year.
    Returns (section, year) or None if not a target file.
    """
    fl = fname.lower()

    # Skip non-spreadsheet files
    ext = os.path.splitext(fname)[1].lower()
    if ext not in (".xlsx", ".xls", ".ods", ".xlsm"):
        return None

    # Skip staff data
    if "staff" in fl:
        return None

    year = extract_year_from_filename(fname)

    # Section 15: attrition/retention/success
    if "section15" in fl.replace(" ", "").replace("_", "") or "section_15" in fl:
        return ("section-15", year or 0)
    if "attrition" in fl and "success" in fl:
        return ("section-15", year or 0)
    if "attrition" in fl and "retention" in fl:
        return ("section-15", year or 0)
    if re.search(r"appendix.?4.*attrition", fl):
        return ("section-15", year or 0)

    # Section 17: completion rates
    if "section17" in fl.replace(" ", "").replace("_", "") or "section_17" in fl:
        return ("section-17", year or 0)
    if "completion rate" in fl and "cohort" not in fl:
        return ("section-17", year or 0)

    # Cohort analysis
    if "cohort" in fl and "completion" in fl:
        return ("cohort-analysis", year or 0)

    # Section 2: all students
    if "section2" in fl.replace(" ", "").replace("_", "").replace("-", "") or "section_2_" in fl:
        return ("section-2", year or 0)
    if "all_students" in fl or "all students" in fl or "all_student" in fl:
        if "first_half" not in fl:
            return ("section-2", year or 0)

    # Section 1: commencing students
    if "section1" in fl.replace(" ", "").replace("_", "").replace("-", "") and "section1" in fl.replace(" ", "").replace("_", "").replace("-", "")[:9]:
        return ("section-1", year or 0)
    if ("commencing_students" in fl or "commencing students" in fl or "commencing_student" in fl) and "load" not in fl:
        if "first_half" not in fl:
            return ("section-1", year or 0)

    # Section 14: award completions
    if "section14" in fl.replace(" ", "").replace("_", "") or "section_14" in fl:
        return ("section-14", year or 0)
    if "award" in fl and "completion" in fl and "pivot" not in fl and "cohort" not in fl:
        return ("section-14", year or 0)
    if "awrdcoursecompletions" in fl.replace(" ", "").replace("_", ""):
        return ("section-14", year or 0)

    # Section 16: equity performance data
    if "section16" in fl.replace(" ", "").replace("_", "") or "section_16" in fl:
        return ("section-16", year or 0)
    if "equity" in fl and "performance" in fl:
        return ("section-16", year or 0)

    # Pivot tables (perturbed and non-perturbed)
    if "pivot" in fl:
        if "completion" in fl or "award" in fl:
            return ("pivot-completions", year or 0)
        elif "enrolment" in fl:
            return ("pivot-enrolments", year or 0)
        elif "load" in fl:
            return ("pivot-load", year or 0)

    return None


def ingest_all(db_path: str, data_dir: str, verbose: bool = False) -> None:
    """Main ingestion entry point."""
    conn = sqlite3.Connection(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")

    # Create schema
    conn.executescript(SCHEMA_DDL)
    conn.commit()

    registry = InstitutionRegistry(conn)
    fields = FieldRegistry(conn)
    fields.seed()
    conn.commit()

    # Scan and classify files
    all_files = sorted(os.listdir(data_dir))
    classified: Dict[str, List[Tuple[str, int]]] = {}

    for fname in all_files:
        result = classify_file(fname)
        if result:
            section, year = result
            classified.setdefault(section, []).append((fname, year))

    print(f"[INFO] Files classified for ingestion:")
    for section, files in sorted(classified.items()):
        print(f"  {section}: {len(files)} files")
    print()

    total_rows = 0

    # Process Section 15 (attrition/retention/success)
    for fname, year in sorted(classified.get("section-15", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            if verbose:
                print(f"  [SKIP] Already ingested: {fname}")
            continue
        print(f"  [S15] Parsing: {fname}")
        try:
            rows = parse_section_15(fpath, conn, registry)
            record_ingestion(conn, fname, fpath, rows, "section-15", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Process Section 17 (completion rates)
    for fname, year in sorted(classified.get("section-17", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            continue
        print(f"  [S17] Parsing: {fname}")
        try:
            rows = parse_section_17(fpath, conn, registry)
            record_ingestion(conn, fname, fpath, rows, "section-17", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Process Cohort Analysis
    for fname, year in sorted(classified.get("cohort-analysis", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            continue
        print(f"  [COH] Parsing: {fname}")
        try:
            rows = parse_cohort_analysis(fpath, conn, registry)
            record_ingestion(conn, fname, fpath, rows, "cohort-analysis", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Process Section 2 (all students enrolments)
    for fname, year in sorted(classified.get("section-2", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            continue
        if not year:
            continue
        print(f"  [S02] Parsing: {fname}")
        try:
            rows = parse_enrolments(fpath, conn, registry, fields, year, commencing=0)
            record_ingestion(conn, fname, fpath, rows, "section-2", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Process Section 1 (commencing students)
    for fname, year in sorted(classified.get("section-1", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            continue
        if not year:
            continue
        print(f"  [S01] Parsing: {fname}")
        try:
            rows = parse_enrolments(fpath, conn, registry, fields, year, commencing=1)
            record_ingestion(conn, fname, fpath, rows, "section-1", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Process Section 14 (award completions)
    for fname, year in sorted(classified.get("section-14", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            continue
        if not year:
            continue
        print(f"  [S14] Parsing: {fname}")
        try:
            rows = parse_section_14(fpath, conn, registry, fields, year)
            record_ingestion(conn, fname, fpath, rows, "section-14", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Process Pivot Tables
    for section_key, is_comp in [("pivot-enrolments", False), ("pivot-completions", True)]:
        for fname, year in sorted(classified.get(section_key, []), key=lambda x: x[1]):
            fpath = os.path.join(data_dir, fname)
            if already_ingested(conn, fname):
                continue
            label = "PIV-C" if is_comp else "PIV-E"
            print(f"  [{label}] Parsing: {fname}")
            try:
                rows = parse_pivot_table(fpath, conn, registry, fields, is_completions=is_comp)
                record_ingestion(conn, fname, fpath, rows, section_key, str(year))
                conn.commit()
                total_rows += rows
                print(f"         -> {rows} rows")
            except Exception as e:
                print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
                conn.rollback()

    # Process Section 16 (equity performance data)
    for fname, year in sorted(classified.get("section-16", []), key=lambda x: x[1]):
        fpath = os.path.join(data_dir, fname)
        if already_ingested(conn, fname):
            if verbose:
                print(f"  [SKIP] Already ingested: {fname}")
            continue
        print(f"  [S16] Parsing: {fname}")
        try:
            rows = parse_section_16(fpath, conn, registry)
            record_ingestion(conn, fname, fpath, rows, "section-16", str(year))
            conn.commit()
            total_rows += rows
            print(f"         -> {rows} rows")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}", file=sys.stderr)
            conn.rollback()

    # Final summary
    print(f"\n{'='*60}")
    print(f"[INFO] Ingestion complete. Total rows inserted: {total_rows}")
    print(f"[INFO] Database: {db_path}")

    for table in ["institutions", "fields_of_education", "attrition_retention",
                   "completion_rates", "enrolments", "completions",
                   "equity_performance", "ingested_files"]:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table}: {count:,} rows")

    conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)

    ap = argparse.ArgumentParser(
        description="Ingest Higher Education Statistics files into SQLite"
    )
    ap.add_argument("--db", type=str, default="he_stats.db", help="SQLite database path")
    ap.add_argument("--data-dir", type=str, default="_downloads/files", help="Directory with downloaded files")
    ap.add_argument("--verbose", action="store_true", help="Show skip messages")
    args = ap.parse_args()

    if not os.path.isdir(args.data_dir):
        print(f"[ERROR] Data directory not found: {args.data_dir}", file=sys.stderr)
        sys.exit(1)

    ingest_all(args.db, args.data_dir, verbose=args.verbose)


if __name__ == "__main__":
    main()
