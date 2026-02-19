"""
Ingest student-staff ratios from the 2024 Staff A2 publication.

Source: 2024 Staff A2 Student Staff Ratios_updated.xlsx
  - Sheet A2.1: Student-Staff Ratios (EFTSL / FTE), 2014-2023
  - Sheet A2.2: Raw EFTSL and Staff FTE, 2014-2023

Creates and populates the `student_staff_ratios` table with:
  - Academic staff ratio (students per academic FTE)
  - Non-academic staff ratio (students per non-academic FTE)
  - Raw EFTSL, academic FTE, and non-academic FTE numbers

Run once:
    python3.9 ingest_staff_ratios.py
"""

from pathlib import Path
import sqlite3
import openpyxl

DB_PATH = Path(__file__).resolve().parent / "he_stats.db"
FILE = Path(__file__).resolve().parent / "_downloads" / "files" / "2024 Staff A2 Student Staff Ratios_updated.xlsx"

SOURCE_FILE = "2024 Staff A2 Student Staff Ratios_updated.xlsx"


def main():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")

    # Create table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS student_staff_ratios (
            id                  INTEGER PRIMARY KEY,
            institution_id      INTEGER NOT NULL REFERENCES institutions(id),
            year                INTEGER NOT NULL,
            academic_ratio      REAL,      -- EFTSL per academic FTE (incl casual)
            non_academic_ratio  REAL,      -- EFTSL per non-academic FTE (incl casual)
            eftsl               REAL,      -- onshore student EFTSL
            academic_fte        REAL,      -- academic staff FTE (incl casual)
            non_academic_fte    REAL,      -- non-academic staff FTE (incl casual)
            source_file         TEXT,
            UNIQUE(institution_id, year)
        )
    """)

    # Build institution code -> id lookup
    rows = conn.execute(
        "SELECT id, code FROM institutions WHERE code IS NOT NULL"
    ).fetchall()
    code_to_id = {r["code"]: r["id"] for r in rows}

    # Also build name -> id lookup (for fallback matching)
    name_rows = conn.execute(
        "SELECT id, name FROM institutions"
    ).fetchall()
    # Normalise names for matching
    name_to_id = {}
    for r in name_rows:
        norm = r["name"].strip().lower()
        name_to_id[norm] = r["id"]

    # Also include aliases
    alias_rows = conn.execute(
        "SELECT alias, institution_id FROM institution_aliases"
    ).fetchall()
    for r in alias_rows:
        name_to_id[r["alias"].strip().lower()] = r["institution_id"]

    wb = openpyxl.load_workbook(str(FILE), data_only=True)

    # ── Parse A2.1: Ratios ──────────────────────────────────────────
    ws_ratios = wb["A2.1"]

    # Year columns: row 5, cols 5-14 for academic (2014-2023), cols 16-25 for non-academic
    year_row = list(ws_ratios.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    academic_years = [int(year_row[i]) for i in range(4, 14)]     # cols E-N (0-indexed: 4-13)
    non_acad_years = [int(year_row[i]) for i in range(15, 25)]    # cols P-Y (0-indexed: 15-24)

    assert academic_years == non_acad_years, "Year columns mismatch!"
    years = academic_years

    # Parse institution rows (starting at row 6)
    ratios = {}  # (inst_id, year) -> {academic_ratio, non_academic_ratio}

    for row in ws_ratios.iter_rows(min_row=6, max_row=60, values_only=False):
        code_val = row[1].value  # Col B = institution code
        name_val = row[2].value  # Col C = institution name

        if not code_val or not name_val:
            continue
        if "total" in str(name_val).lower() or "note" in str(name_val).lower():
            continue

        code_str = str(int(code_val))
        inst_id = code_to_id.get(code_str)

        # Fallback: name matching
        if inst_id is None:
            norm_name = name_val.strip().lower()
            inst_id = name_to_id.get(norm_name)

        if inst_id is None:
            print(f"  WARN: No match for code={code_str} name={name_val}")
            continue

        for i, yr in enumerate(years):
            acad_ratio = row[4 + i].value     # cols E-N
            non_acad_ratio = row[15 + i].value  # cols P-Y

            ratios[(inst_id, yr)] = {
                "academic_ratio": round(float(acad_ratio), 2) if isinstance(acad_ratio, (int, float)) else None,
                "non_academic_ratio": round(float(non_acad_ratio), 2) if isinstance(non_acad_ratio, (int, float)) else None,
            }

    # ── Parse A2.2: Raw EFTSL and FTE ───────────────────────────────
    ws_raw = wb["A2.2"]

    for row in ws_raw.iter_rows(min_row=6, max_row=60, values_only=False):
        code_val = row[1].value
        name_val = row[2].value

        if not code_val or not name_val:
            continue
        if "total" in str(name_val).lower() or "note" in str(name_val).lower():
            continue

        code_str = str(int(code_val))
        inst_id = code_to_id.get(code_str)
        if inst_id is None:
            norm_name = name_val.strip().lower()
            inst_id = name_to_id.get(norm_name)
        if inst_id is None:
            continue

        for i, yr in enumerate(years):
            key = (inst_id, yr)
            if key not in ratios:
                ratios[key] = {"academic_ratio": None, "non_academic_ratio": None}

            eftsl = row[4 + i].value          # cols E-N: EFTSL
            acad_fte = row[26 + i].value       # cols AA-AJ: Academic FTE
            non_acad_fte = row[37 + i].value   # cols AL-AU: Non-academic FTE

            ratios[key]["eftsl"] = round(float(eftsl), 1) if isinstance(eftsl, (int, float)) else None
            ratios[key]["academic_fte"] = round(float(acad_fte), 1) if isinstance(acad_fte, (int, float)) else None
            ratios[key]["non_academic_fte"] = round(float(non_acad_fte), 1) if isinstance(non_acad_fte, (int, float)) else None

    # ── Insert into DB ──────────────────────────────────────────────
    inserted = 0
    for (inst_id, yr), data in ratios.items():
        conn.execute(
            """INSERT OR REPLACE INTO student_staff_ratios
               (institution_id, year, academic_ratio, non_academic_ratio,
                eftsl, academic_fte, non_academic_fte, source_file)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                inst_id,
                yr,
                data.get("academic_ratio"),
                data.get("non_academic_ratio"),
                data.get("eftsl"),
                data.get("academic_fte"),
                data.get("non_academic_fte"),
                SOURCE_FILE,
            ),
        )
        inserted += 1

    conn.commit()
    total = conn.execute("SELECT COUNT(*) as c FROM student_staff_ratios").fetchone()["c"]
    print(f"Inserted {inserted} rows. Table total: {total}")

    # Quick sanity check
    sample = conn.execute("""
        SELECT i.name, s.year, s.academic_ratio, s.non_academic_ratio, s.eftsl, s.academic_fte
        FROM student_staff_ratios s
        JOIN institutions i ON i.id = s.institution_id
        WHERE s.year = 2023
        ORDER BY s.academic_ratio
        LIMIT 5
    """).fetchall()
    print("\nLowest academic ratios (2023):")
    for r in sample:
        print(f"  {r['name']}: {r['academic_ratio']} students/academic FTE (EFTSL={r['eftsl']}, FTE={r['academic_fte']})")

    sample2 = conn.execute("""
        SELECT i.name, s.year, s.academic_ratio
        FROM student_staff_ratios s
        JOIN institutions i ON i.id = s.institution_id
        WHERE s.year = 2023
        ORDER BY s.academic_ratio DESC
        LIMIT 5
    """).fetchall()
    print("\nHighest academic ratios (2023):")
    for r in sample2:
        print(f"  {r['name']}: {r['academic_ratio']} students/academic FTE")

    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
